import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter



def daily_download(project_name,line_id, camera_ip, vehicle_counts,vehicle_counts_up,vehicle_counts_down,vehicle_classes, start_date, end_date, 
                   line_1_data, line_2_data, line_3_up_data, line_3_down_data, output,resolved_line_ids,polygon_1_data,polygon_2_data,polygon_3_data):
    
    print("start_date:",start_date,end_date)
    
    print("daily downloaded.............................",line_1_data)

    start_date = pd.to_datetime(start_date)
    end_date = pd.to_datetime(end_date)
    end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=0)

    print("cureent start date end date:",start_date,end_date)

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        current_date = start_date

        print("##############current date:#################:",current_date,end_date)
        while current_date <= end_date:
            print("#$$$$$$$$$$$$$$$$$@##########$#%#^%********************")
            sheet_name = current_date.strftime('%d%b%y')
            pd.DataFrame().to_excel(writer, index=False, sheet_name=sheet_name, startrow=5)

            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            bold_side = Side(style='thin')
            thin_border = Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)

            for row in worksheet.iter_rows():
                for cell in row:
                    cell.font = Font(name='Aptos Narrow')
                    cell.alignment = Alignment(horizontal='center')  
                    cell.border = Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)

            worksheet['A1'] = f'ATCC Day Report'
            worksheet['A2'] = f'Daily Report for {sheet_name}'
            worksheet['A3'] = 'Report generated: '
            worksheet['C3'] = datetime.now().strftime('%d%b%y')
            worksheet['D3'] = f'Project:{project_name}-{camera_ip}'
            worksheet['D4'] = "line_1"
           
            worksheet['A31'] = 'Overall Total'

            for cell in ['A1', 'A2', 'A3', 'C3', 'D3', 'D4','A31']:
                worksheet[cell].border = thin_border

            centered_cells = ['A1', 'A2', 'A3', 'C3', 'D3', 'D4']
            for cell in centered_cells:
                worksheet[cell].alignment = Alignment(horizontal='center')


            vehicle_class_keys = list(vehicle_classes.keys())
            
            colors = {
                'title': "660066",
                'header': "FF0000",
                'sub_header': "00CCFF"
            }
            
            
            
            worksheet.merge_cells(start_row=3, start_column=1, end_row=5, end_column=2)
            worksheet.merge_cells(start_row=3, start_column=3, end_row=5, end_column=3)

            worksheet.merge_cells(start_row=31, start_column=1, end_row=31, end_column=3)

            lhs_start_col = 4
            lhs_end_col = lhs_start_col + len(vehicle_class_keys)
            lhs_count_start_col = 4
            rhs_count_start_col = lhs_count_start_col + len(vehicle_class_keys) + 1  
            total_count_start_col = rhs_count_start_col + len(vehicle_class_keys) + 1
            if line_id.startswith("line_"):
                for lid in resolved_line_ids:
                    print("lid:",lid)
                
                    if 'line_id_1' in resolved_line_ids and 'line_id_2' in resolved_line_ids:
                                print("helloooooooooooooo kishs",1111111111111111)
                                lhs_start_col=4
                                print("line_id_1 and line_id_2",lhs_start_col)
                            

                                up_cell = worksheet.cell(row=5, column=lhs_start_col)
                                up_cell.value = "UP"
                                up_cell.font = Font(name='Aptos Narrow',size=10, bold=True, color="00CCFF")  
                                up_cell.alignment = Alignment(horizontal='center')  
                                up_cell.border=Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)

                                vehicle_class_keys = list(vehicle_classes.keys())
                                headers = ['Date', 'Start Time', 'End Time', 
                                            'Total Vehicles (UP)', *[(f'{cls} ') for cls in vehicle_class_keys]]
                                
                                for i, header in enumerate(headers, start=1):
                                    cell = worksheet.cell(row=6, column=i, value=header)
                                    cell.font = Font(name='Aptos Narrow',size=10,bold=True)
                                    cell.alignment = Alignment(horizontal='center', wrap_text=True) 
                                    cell.border = Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)
                                
                                last_column_index = len(headers) 
                                worksheet.merge_cells(start_row=4, start_column=lhs_start_col, end_row=4, end_column=last_column_index)
                                worksheet.merge_cells(start_row=5, start_column=lhs_start_col, end_row=5, end_column=last_column_index)

                    

                            
                                
                                start_row = 7
                                print("hourly vehicle counts:", vehicle_counts)

                                if str(current_date.date()) in vehicle_counts:
                                    print("!!!!!!!!!!!!!!!!!!!!!!")
                                    hours = vehicle_counts[str(current_date.date())]
                                    print("hours:", hours)

                                grand_total_lhs = 0  

                                for hour in range(24):
                                    print(f"111111Processing hour: {hour}")
                                    total_lhs = 0
                                    counts_lhs = {cls: 0 for cls in vehicle_class_keys}
                                    
                                    #####kishore
                                    date_key = str(current_date.date())  # "2025-09-17"

                                    if date_key in line_1_data['line_id_1']:
                                        print(f"Processing line1 data: {line_1_data}")
                                        kit1_counts = line_1_data['line_id_1'][date_key].get(hour, {cls: 0 for cls in vehicle_class_keys})
                                        print("kit1 counts:", kit1_counts)

                                        for cls in vehicle_class_keys:
                                            counts_lhs[cls] += kit1_counts.get(cls, 0)
                                            total_lhs += counts_lhs[cls]
                                            print("lhs:", counts_lhs[cls])
                                    else:
                                        print("nooooo else 11111111111111", current_date)

                                    ##kishore

                                    # if str(current_date) in line_1_data:
                                    #     print(f"Processing line1 data: {line_1_data}")
                                    #     kit1_counts = line_1_data[str(current_date.date())].get(hour, {cls: 0 for cls in vehicle_class_keys})
                                    #     print("kit1 counts:", kit1_counts)

                                    #     for cls in vehicle_class_keys:
                                    #         counts_lhs[cls] += kit1_counts.get(cls, 0)
                                    #         total_lhs += counts_lhs[cls]
                                    #         print("lhs:", counts_lhs[cls])
                                    # else:
                                    #     print("noooooooooooooo else 11111111111111",current_date)


                                    start_time = f"{hour:02d}:00"
                                    end_time = f"{hour + 1:02d}:00" if hour < 23 else "23:59"

                                    print("start_time:", start_time)
                                    print("end_time:", end_time)

                                    
                                    total_cell = worksheet.cell(row=start_row, column=4, value=total_lhs)  
                                    total_cell.border = thin_border  
                                    total_cell.alignment = Alignment(horizontal='center', wrap_text=True)

                                    worksheet.cell(row=start_row, column=1, value=current_date.strftime('%d%b%y').lower()).border = thin_border
                                    worksheet.cell(row=start_row, column=2, value=start_time).border = thin_border
                                    worksheet.cell(row=start_row, column=3, value=end_time).border = thin_border


                                    for i, cls in enumerate(vehicle_class_keys):
                                        print("#R$#%$#%$#%")
                                        lhs_cell = worksheet.cell(row=start_row, column=4 + i + 1, value=counts_lhs[cls])
                                        lhs_cell.border = thin_border   
                                        lhs_cell.alignment = Alignment(horizontal='center', wrap_text=True) 
                                        lhs_cell.font = Font(name='Aptos Narrow', size=10)

                                    grand_total_lhs += total_lhs  
                                    start_row += 1  

                        

                                grand_total_cell = worksheet.cell(row=start_row, column=4, value=grand_total_lhs)  
                                grand_total_cell.alignment = Alignment(horizontal='center', wrap_text=True)
                                grand_total_cell.border = thin_border  
                                # last_column_index = rhs_end_col

                                overall_totals = [0] * (len(headers)+3)
                                for row in range(7, start_row):
                                    for col in range(4, len(headers) + 3):
                                        overall_totals[col - 1] += worksheet.cell(row=row, column=col).value or 0

                                        for col in range(4, len(headers) + 3):
                                            total_cell=worksheet.cell(row=start_row, column=col, value=overall_totals[col - 1])
                                            total_cell.border = thin_border
                                            total_cell.alignment = Alignment(horizontal='center') 
                                            total_cell.font = Font(name='Aptos Narrow', size=10)
                                #current_date += timedelta(days=1)
                                print("last column index:",last_column_index)
                                lhs_start_col = last_column_index + 1  
                                total_start_col = lhs_start_col
                                total_end_col = total_start_col + len(vehicle_class_keys)

                                print("lhs start,total_start,total_end:",lhs_start_col,total_start_col,total_end_col)
                                
                                print("hourly vehicle counts:",vehicle_counts)

                                ######LINE_ID_2###-----------------------------------------------

                                if str(current_date.date()) in vehicle_counts:
                                    print("!!!!!!!!!!!!!!!!!!!!!!")
                                    hours = vehicle_counts[str(current_date.date())]
                                    print("hours:",hours)     


                                for hour in range(24):
                                    print(f"Processing hour: {hour}")
                                    total_lhs = total_rhs = total_counts = 0
                                    counts_lhs = {cls: 0 for cls in vehicle_class_keys}
                                    counts_rhs = {cls: 0 for cls in vehicle_class_keys}
                
                                down_cell = worksheet.cell(row=5, column=lhs_start_col)
                                down_cell.value = "DOWN"
                                down_cell.font = Font(name='Aptos Narrow',size=10, bold=True, color="00CCFF")  
                                down_cell.alignment = Alignment(horizontal='center')  
                                down_cell.border=Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)

                                towards_padina_cell = worksheet.cell(row=4, column=lhs_start_col)
                                towards_padina_cell.value = "line_2"
                                towards_padina_cell.font = Font(name='Aptos Narrow',size=10, bold=True, color="00CCFF")  
                                towards_padina_cell.alignment = Alignment(horizontal='center')  
                                towards_padina_cell.border=Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)

                            
                                total_heading_cell = worksheet.cell(row=3, column=lhs_start_col)

                                vehicle_class_keys = list(vehicle_classes.keys())
                                headers = ['Date', 'Start Time', 'End Time', 
                                            'Total Vehicles (UP)', *[(f'{cls} ') for cls in vehicle_class_keys],
                                        
                                    'Total Vehicles (DOWN)', *[(f'{cls} ') for cls in vehicle_class_keys]]

                                for i, header in enumerate(headers, start=1):
                                    print("i",i)
                                    cell = worksheet.cell(row=6, column=i, value=header)
                                    cell.font = Font(name='Aptos Narrow',size=10,bold=True)
                                    cell.alignment = Alignment(horizontal='center',wrap_text=True) 
                                    cell.border = Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)

                                print("lhs start,total_start,total_end:",lhs_start_col,total_start_col,total_end_col,last_column_index)
                                worksheet.merge_cells(start_row=3, start_column=4, end_row=3, end_column=last_column_index+len(vehicle_classes)+1)


                                worksheet.merge_cells(start_row=4, start_column=lhs_start_col, end_row=4, end_column=last_column_index+len(vehicle_classes)+1)
                                worksheet.merge_cells(start_row=5, start_column=lhs_start_col, end_row=5, end_column=last_column_index+len(vehicle_classes)+1)

                    
                                worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column_index+len(vehicle_classes)+1)
                                worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column_index+len(vehicle_classes)+1)


                                print("current date:", current_date)
                                print("Available keys in line_2_data:", line_2_data['line_id_2'].keys())
                                print("Formatted current date:", str(current_date.date()))

                                print("Type of current_date.date():", type(current_date.date()))
                                print("Type of keys in line_2_data['line_id_2']:", type(list(line_2_data['line_id_2'].keys())[0]))

                                formatted_date = str(current_date.date())

                                print("formatted_date in line id 2 :",formatted_date)

                                total_rhs_list = []  
                                row_number = 7  

                                if formatted_date in line_2_data['line_id_2']:
                                    print(f"Processing line2 data:{line_2_data}")

                                    for hour in range(24):  
                                    
                                        kit2_counts = line_2_data['line_id_2'][formatted_date].get(hour, {cls: 0 for cls in vehicle_class_keys})
                                        counts_rhs = {cls: 0 for cls in vehicle_class_keys}  

                                        for cls in vehicle_class_keys:
                                            counts_rhs[cls] = kit2_counts.get(cls, 0)

                                        total_rhs = sum(counts_rhs.values())  
                                        total_rhs_list.append(total_rhs)  

                                        print(f"Hour: {hour}, Total RHS Count: {total_rhs}")

                                        start_time = f"{hour:02d}:00"
                                        end_time = f"{hour + 1:02d}:00" if hour < 23 else "23:59"

                                        print("start_time:", start_time)
                                        print("end_time:", end_time)

                                        for i, cls in enumerate(vehicle_class_keys):
                                            rhs_cell = worksheet.cell(row=row_number, column=total_start_col+ i + 1, value=counts_rhs[cls])
                                            rhs_cell.border = thin_border
                                            rhs_cell.alignment = Alignment(horizontal='center')
                                            rhs_cell.font = Font(name='Aptos Narrow', size=10)

                                    
                                        total_cell = worksheet.cell(row=row_number, column=total_start_col, value=total_rhs)
                                        total_cell.border = thin_border
                                        total_cell.alignment = Alignment(horizontal='center')
                                        total_cell.font = Font(name='Aptos Narrow', size=10)

                                        start_row += 1 
                                        row_number += 1  

                                    grand_total = sum(total_rhs_list)
                                    grand_total_cell = worksheet.cell(row=row_number, column=total_start_col, value=grand_total)
                                    grand_total_cell.border = thin_border
                                    grand_total_cell.alignment = Alignment(horizontal='center')
                                    grand_total_cell.font = Font(name='Aptos Narrow', size=10) 

                                    print(f"Grand Total RHS Count: {grand_total}")

                                    overall_totals = [0] * (len(headers)) 
                                    print("overall totals:", overall_totals, len(headers) + 1)


                                    for row in range(7, start_row):
                                        for col in range(4, len(headers)):  
                                            overall_totals[col - 4] += worksheet.cell(row=row, column=col).value or 0 

                                
                                    for col in range(4, len(headers) ): 
                                        total_cell = worksheet.cell(row=start_row, column=col, value=overall_totals[col - 4])  
                                        total_cell.border = thin_border
                                        total_cell.alignment = Alignment(horizontal='center') 
                                        total_cell.font = Font(name='Aptos Narrow', size=10)

                                # current_date += timedelta(days=1)



                    elif lid =='line_id_1':
                        print("line id 1 ")
                
                    
                        rhs_start_col = lhs_end_col + 1  
                        rhs_end_col = rhs_start_col + len(vehicle_class_keys)

                        up_cell = worksheet.cell(row=5, column=lhs_start_col)
                        up_cell.value = "UP"
                        up_cell.font = Font(name='Aptos Narrow',size=10, bold=True, color="00CCFF")  
                        up_cell.alignment = Alignment(horizontal='center')  
                        up_cell.border=Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)

                        vehicle_class_keys = list(vehicle_classes.keys())
                        headers = ['Date', 'Start Time', 'End Time', 
                        'Total Vehicles (UP)', *[(f'{cls} ') for cls in vehicle_class_keys]]

                        for i, header in enumerate(headers, start=1):
                            cell = worksheet.cell(row=6, column=i, value=header)
                            cell.font = Font(name='Aptos Narrow',size=10,bold=True)
                            cell.alignment = Alignment(horizontal='center', wrap_text=True) 
                            cell.border = Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)
                        
                        last_column_index = len(headers) 
                        worksheet.merge_cells(start_row=4, start_column=lhs_start_col, end_row=4, end_column=last_column_index)
                        worksheet.merge_cells(start_row=5, start_column=lhs_start_col, end_row=5, end_column=last_column_index)

            

                        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column_index)
                        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column_index)
                        
                        start_row = 7
                        print("hourly vehicle counts:", vehicle_counts)

                        if str(current_date.date()) in vehicle_counts:
                            print("!!!!!!!!!!!!!!!!!!!!!!")
                            hours = vehicle_counts[str(current_date.date())]
                            print("hours:", hours)

                            grand_total_lhs = 0  

                            for hour in range(24):
                                print(f"Processing hour: {hour}")
                                total_lhs = 0
                                counts_lhs = {cls: 0 for cls in vehicle_class_keys}
                                
                                print("cureecnt date:",current_date.date(),line_1_data)

                                if 'line_id_1' in line_1_data and str(current_date.date()) in line_1_data['line_id_1']:
                                

                                    print(f"Processing line1 data: {line_1_data}")
                                    kit1_counts = line_1_data['line_id_1'][str(current_date.date())].get(hour, {cls: 0 for cls in vehicle_class_keys})
                                    print("kit1 counts:", kit1_counts)

                                    for cls in vehicle_class_keys:
                                        counts_lhs[cls] += kit1_counts.get(cls, 0)
                                        total_lhs += counts_lhs[cls]
                                        print("lhs:", counts_lhs[cls])

                                start_time = f"{hour:02d}:00"
                                end_time = f"{hour + 1:02d}:00" if hour < 23 else "23:59"

                                print("start_time in line id 1:", start_time)
                                print("end_time:", end_time)

                                
                                total_cell = worksheet.cell(row=start_row, column=4, value=total_lhs)  
                                total_cell.border = thin_border  
                                total_cell.alignment = Alignment(horizontal='center', wrap_text=True)

                                worksheet.cell(row=start_row, column=1, value=current_date.strftime('%d%b%y').lower()).border = thin_border
                                worksheet.cell(row=start_row, column=2, value=start_time).border = thin_border
                                worksheet.cell(row=start_row, column=3, value=end_time).border = thin_border
                                worksheet.merge_cells(start_row=3, start_column=lhs_start_col, end_row=3, end_column=last_column_index)


                                for i, cls in enumerate(vehicle_class_keys):

                                    lhs_cell = worksheet.cell(row=start_row, column=4 + i + 1, value=counts_lhs[cls])
                                    lhs_cell.border = thin_border   
                                    lhs_cell.alignment = Alignment(horizontal='center', wrap_text=True) 
                                    lhs_cell.font = Font(name='Aptos Narrow', size=10)

                                grand_total_lhs += total_lhs  
                                start_row += 1  

                    

                            grand_total_cell = worksheet.cell(row=start_row, column=4, value=grand_total_lhs)  
                            grand_total_cell.alignment = Alignment(horizontal='center', wrap_text=True)
                            grand_total_cell.border = thin_border  
                            

                            overall_totals = [0] * len(headers)
                            for row in range(7, start_row):
                                for col in range(4, len(headers) + 1):
                                    overall_totals[col - 1] += worksheet.cell(row=row, column=col).value or 0

                                    for col in range(4, len(headers) + 1):
                                        total_cell=worksheet.cell(row=start_row, column=col, value=overall_totals[col - 1])
                                        total_cell.border = thin_border
                                        total_cell.alignment = Alignment(horizontal='center') 
                                        total_cell.font = Font(name='Aptos Narrow', size=10)
                        # current_date += timedelta(days=1)

                        

                    elif lid == 'line_id_2':
                        print("line id 2")
                        start_row = 7
                        print("hourly vehicle counts:",vehicle_counts)

                        if str(current_date.date()) in vehicle_counts:
                            print("!!!!!!!!!!!!!!!!!!!!!!")
                            hours = vehicle_counts[str(current_date.date())]
                            print("hours:",hours)     


                            for hour in range(24):
                                print(f"Processing hour: {hour}")
                                total_lhs = total_rhs = total_counts = 0
                                counts_lhs = {cls: 0 for cls in vehicle_class_keys}
                                counts_rhs = {cls: 0 for cls in vehicle_class_keys}
                    #
                        
                            total_start_col = lhs_end_col + 1 
                            total_end_col = total_start_col + len(vehicle_class_keys) 

            

                            down_cell = worksheet.cell(row=5, column=lhs_start_col)
                            down_cell.value = "DOWN"
                            down_cell.font = Font(name='Aptos Narrow',size=10, bold=True, color="00CCFF")  
                            down_cell.alignment = Alignment(horizontal='center')  
                            down_cell.border=Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)



                            towards_padina_cell = worksheet.cell(row=4, column=lhs_start_col)
                            towards_padina_cell.value = "line_2"
                            towards_padina_cell.font = Font(name='Aptos Narrow',size=10, bold=True, color="00CCFF")  
                            towards_padina_cell.alignment = Alignment(horizontal='center')  
                            towards_padina_cell.border=Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)

                        
                            total_heading_cell = worksheet.cell(row=3, column=lhs_start_col)

                            vehicle_class_keys = list(vehicle_classes.keys())
                            headers = ['Date', 'Start Time', 'End Time', 
                                
                            'Total Vehicles (DOWN)', *[(f'{cls} ') for cls in vehicle_class_keys]]

                            for i, header in enumerate(headers, start=1):
                                cell = worksheet.cell(row=6, column=i, value=header)
                                cell.font = Font(name='Aptos Narrow',size=10,bold=True)
                                cell.alignment = Alignment(horizontal='center',wrap_text=True) 
                                cell.border = Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)

                            last_column_index = len(headers) 
                            worksheet.merge_cells(start_row=3, start_column=lhs_start_col, end_row=3, end_column=last_column_index)


                            worksheet.merge_cells(start_row=4, start_column=lhs_start_col, end_row=4, end_column=last_column_index)
                            worksheet.merge_cells(start_row=5, start_column=lhs_start_col, end_row=5, end_column=last_column_index)

                            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column_index)
                            worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column_index)


                            print("current date:", current_date)
                            print("Available keys in line_2_data:", line_2_data['line_id_2'].keys())
                            print("Formatted current date:", str(current_date.date()))

                            print("Type of current_date.date():", type(current_date.date()))
                            print("Type of keys in line_2_data['line_id_2']:", type(list(line_2_data['line_id_2'].keys())[0]))

                            formatted_date = str(current_date.date())

                            total_rhs_list = []  
                            row_number = 7  

                            if formatted_date in line_2_data['line_id_2']:
                                print(f"Processing line2 data:{line_2_data}")

                                for hour in range(24):  
                                
                                    kit2_counts = line_2_data['line_id_2'][formatted_date].get(hour, {cls: 0 for cls in vehicle_class_keys})
                                    counts_rhs = {cls: 0 for cls in vehicle_class_keys}  

                                    for cls in vehicle_class_keys:
                                        counts_rhs[cls] = kit2_counts.get(cls, 0)

                                    total_rhs = sum(counts_rhs.values())  
                                    total_rhs_list.append(total_rhs)  

                                    print(f"Hour: {hour}, Total RHS Count: {total_rhs}")

                                    start_time = f"{hour:02d}:00"
                                    end_time = f"{hour + 1:02d}:00" if hour < 23 else "23:59"

                                    print("start_time:", start_time)
                                    print("end_time:", end_time)

                                
                                    worksheet.cell(row=start_row, column=1, value=current_date.strftime('%d%b%y').lower()).border = thin_border
                                    worksheet.cell(row=start_row, column=2, value=start_time).border = thin_border
                                    worksheet.cell(row=start_row, column=3, value=end_time).border = thin_border

                                    for i, cls in enumerate(vehicle_class_keys):
                                        rhs_cell = worksheet.cell(row=start_row, column=4 + i + 1, value=counts_rhs[cls])
                                        rhs_cell.border = thin_border
                                        rhs_cell.alignment = Alignment(horizontal='center')
                                        rhs_cell.font = Font(name='Aptos Narrow', size=10)

                                
                                    total_cell = worksheet.cell(row=row_number, column=4, value=total_rhs)
                                    total_cell.border = thin_border
                                    total_cell.alignment = Alignment(horizontal='center')
                                    total_cell.font = Font(name='Aptos Narrow', size=10)

                                    start_row += 1 
                                    row_number += 1  

                                grand_total = sum(total_rhs_list)
                                grand_total_cell = worksheet.cell(row=row_number, column=4, value=grand_total)
                                grand_total_cell.border = thin_border
                                grand_total_cell.alignment = Alignment(horizontal='center')
                                grand_total_cell.font = Font(name='Aptos Narrow', size=10, color="FF0000") 

                                print(f"Grand Total RHS Count: {grand_total}")

                                overall_totals = [0] * len(headers)
                                for row in range(7, start_row):
                                    for col in range(4, len(headers) + 1):
                                        overall_totals[col - 1] += worksheet.cell(row=row, column=col).value or 0


                                        for col in range(4, len(headers) + 1):
                                            total_cell=worksheet.cell(row=start_row, column=col, value=overall_totals[col - 1])
                                            total_cell.border = thin_border
                                            total_cell.alignment = Alignment(horizontal='center') 
                                            total_cell.font = Font(name='Aptos Narrow', size=10)
                        # current_date += timedelta(days=1)

                    elif lid == 'line_id_3_up':  
                        print("line id 3")
                        start_row = 7
                        print("hourly vehicle counts:",vehicle_counts)

                        if str(current_date.date()) in vehicle_counts:
                            print("!!!!!!!!!!!!!!!!!!!!!!")
                            hours = vehicle_counts[str(current_date.date())]
                            print("hours:",hours)     


                        for hour in range(24):
                            print(f"Processing hour: {hour}")
                            total_lhs = total_rhs = total_counts = 0
                            counts_lhs = {cls: 0 for cls in vehicle_class_keys}
                            counts_rhs = {cls: 0 for cls in vehicle_class_keys}
                #
                    
                        total_start_col = lhs_end_col + 1 
                        total_end_col = total_start_col + len(vehicle_class_keys) 

                        down_cell = worksheet.cell(row=5, column=lhs_start_col)
                        down_cell.value = "UP"
                        down_cell.font = Font(name='Aptos Narrow',size=10, bold=True, color="00CCFF")  
                        down_cell.alignment = Alignment(horizontal='center')  
                        down_cell.border=Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)

                        towards_padina_cell = worksheet.cell(row=4, column=lhs_start_col)
                        towards_padina_cell.value = "line_3"
                        towards_padina_cell.font = Font(name='Aptos Narrow',size=10, bold=True, color="00CCFF")  
                        towards_padina_cell.alignment = Alignment(horizontal='center')  
                        towards_padina_cell.border=Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)

                        total_heading_cell = worksheet.cell(row=3, column=lhs_start_col)

                        vehicle_class_keys = list(vehicle_classes.keys())
                        headers = ['Date', 'Start Time', 'End Time', 
                                'Total Vehicles (UP)', *[(f'{cls} ') for cls in vehicle_class_keys],
                            
                        'Total Vehicles (DOWN)', *[(f'{cls} ') for cls in vehicle_class_keys]]

                        for i, header in enumerate(headers, start=1):
                            cell = worksheet.cell(row=6, column=i, value=header)
                            cell.font = Font(name='Aptos Narrow',size=10,bold=True)
                            cell.alignment = Alignment(horizontal='center',wrap_text=True) 
                            cell.border = Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)

                        last_column_index = len(headers) 
                        worksheet.merge_cells(start_row=3, start_column=lhs_start_col, end_row=3, end_column=last_column_index)
                        worksheet.merge_cells(start_row=4, start_column=lhs_start_col, end_row=4, end_column=last_column_index)
                        worksheet.merge_cells(start_row=5, start_column=lhs_start_col, end_row=5, end_column=4+len(vehicle_classes))

                        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column_index)
                        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column_index)

                        print("current date:", current_date)
                        print("Available keys in line_3_up_data:", line_3_up_data['line_id_3_up'].keys())
                        print("Formatted current date:", str(current_date.date()))

                        print("Type of current_date.date():", type(current_date.date()))
                        print("Type of keys in line_2_data['line_id_3_up']:", type(list(line_3_up_data['line_id_3_up'].keys())[0]))

                        formatted_date = str(current_date.date())

                        total_rhs_list = []  
                        row_number = 7  

                        if formatted_date in line_3_up_data['line_id_3_up']:
                            print(f"Processing line2 data:{line_3_up_data}")

                            for hour in range(24):  
                            
                                kit2_counts = line_3_up_data['line_id_3_up'][formatted_date].get(hour, {cls: 0 for cls in vehicle_class_keys})
                                counts_rhs = {cls: 0 for cls in vehicle_class_keys}  

                                for cls in vehicle_class_keys:
                                    counts_rhs[cls] = kit2_counts.get(cls, 0)

                                total_rhs = sum(counts_rhs.values())  
                                total_rhs_list.append(total_rhs)  

                                print(f"Hour: {hour}, Total RHS Count: {total_rhs}")

                                start_time = f"{hour:02d}:00"
                                end_time = f"{hour + 1:02d}:00" if hour < 23 else "23:59"

                                print("start_time:", start_time)
                                print("end_time:", end_time)

                            
                                worksheet.cell(row=start_row, column=1, value=current_date.strftime('%d%b%y').lower()).border = thin_border
                                worksheet.cell(row=start_row, column=2, value=start_time).border = thin_border
                                worksheet.cell(row=start_row, column=3, value=end_time).border = thin_border

                                for i, cls in enumerate(vehicle_class_keys):
                                    rhs_cell = worksheet.cell(row=start_row, column=4 + i + 1, value=counts_rhs[cls])
                                    rhs_cell.border = thin_border
                                    rhs_cell.alignment = Alignment(horizontal='center')
                                    rhs_cell.font = Font(name='Aptos Narrow', size=10)

                            
                                total_cell = worksheet.cell(row=row_number, column=4, value=total_rhs)
                                total_cell.border = thin_border
                                total_cell.alignment = Alignment(horizontal='center')
                                total_cell.font = Font(name='Aptos Narrow', size=10)

                                start_row += 1 
                                row_number += 1  

                            grand_total = sum(total_rhs_list)
                            grand_total_cell = worksheet.cell(row=row_number, column=4, value=grand_total)
                            grand_total_cell.border = thin_border
                            grand_total_cell.alignment = Alignment(horizontal='center')
                            grand_total_cell.font = Font(name='Aptos Narrow', size=10) 

                            print(f"Grand Total RHS Count: {grand_total}")      

                    elif lid == 'line_id_3_down':  
                        print("line id 3")
                        start_row = 7
                        print("hourly vehicle counts:",vehicle_counts)

                        if str(current_date.date()) in vehicle_counts:
                            print("!!!!!!!!!!!!!!!!!!!!!!")
                            hours = vehicle_counts[str(current_date.date())]
                            print("hours:",hours)     


                            for hour in range(24):
                                print(f"Processing hour: {hour}")
                                total_lhs = total_rhs = total_counts = 0
                                counts_lhs = {cls: 0 for cls in vehicle_class_keys}
                                counts_rhs = {cls: 0 for cls in vehicle_class_keys}
                        #
                            
                            total_start_col = lhs_end_col + 1 
                            total_end_col = total_start_col + len(vehicle_class_keys) 

                        
                            down_cell = worksheet.cell(row=5, column=lhs_start_col+len(vehicle_classes)+1)
                            down_cell.value = "Down"
                            down_cell.font = Font(name='Aptos Narrow',size=10, bold=True, color="00CCFF")  
                            down_cell.alignment = Alignment(horizontal='center')  
                            down_cell.border=Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)

                            total_heading_cell = worksheet.cell(row=3, column=lhs_start_col)

                            vehicle_class_keys = list(vehicle_classes.keys())
                            headers = ['Date', 'Start Time', 'End Time', 
                                        'Total Vehicles (UP)', *[(f'{cls} ') for cls in vehicle_class_keys],
                                    
                                'Total Vehicles (DOWN)', *[(f'{cls} ') for cls in vehicle_class_keys]]

                            for i, header in enumerate(headers, start=1):
                                cell = worksheet.cell(row=6, column=i, value=header)
                                cell.font = Font(name='Aptos Narrow',size=10,bold=True)
                                cell.alignment = Alignment(horizontal='center',wrap_text=True) 
                                cell.border = Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)

                            last_column_index = len(headers) 
                            worksheet.merge_cells(start_row=3, start_column=lhs_start_col+len(vehicle_classes)+1, end_row=3, end_column=last_column_index)


                            worksheet.merge_cells(start_row=4, start_column=lhs_start_col+len(vehicle_classes)+1, end_row=4, end_column=last_column_index)
                            worksheet.merge_cells(start_row=5, start_column=lhs_start_col+len(vehicle_classes)+1, end_row=5, end_column=last_column_index)
                                

                            print("current date:", current_date)
                            print("Available keys in line_3_down_data:", line_3_down_data['line_id_3_down'].keys())
                            print("Formatted current date:", str(current_date.date()))

                            print("Type of current_date.date():", type(current_date.date()))
                            print("Type of keys in line_3_down_data['line_id_3_down']:", type(list(line_3_down_data['line_id_3_down'].keys())[0]))

                            formatted_date = str(current_date.date())

                            total_rhs_list = []  
                            row_number = 7  

                            if formatted_date in line_3_down_data['line_id_3_down']:
                                print(f"Processing line3 down data:{line_3_down_data}")

                                for hour in range(24):  
                                    
                                    kit2_counts = line_3_down_data['line_id_3_down'][formatted_date].get(hour, {cls: 0 for cls in vehicle_class_keys})
                                    counts_rhs = {cls: 0 for cls in vehicle_class_keys}  

                                    for cls in vehicle_class_keys:
                                        counts_rhs[cls] = kit2_counts.get(cls, 0)

                                    total_rhs = sum(counts_rhs.values())  
                                    total_rhs_list.append(total_rhs)  

                                    print(f"Hour: {hour}, Total RHS Count: {total_rhs}")

                                    start_time = f"{hour:02d}:00"
                                    end_time = f"{hour + 1:02d}:00" if hour < 23 else "23:59"

                                    print("start_time:", start_time)
                                    print("end_time:", end_time)

                                    
                                    worksheet.cell(row=start_row, column=1, value=current_date.strftime('%d%b%y').lower()).border = thin_border
                                    worksheet.cell(row=start_row, column=2, value=start_time).border = thin_border
                                    worksheet.cell(row=start_row, column=3, value=end_time).border = thin_border

                                    for i, cls in enumerate(vehicle_class_keys):
                                        rhs_cell = worksheet.cell(row=start_row, column=4+len(vehicle_classes) + i + 2, value=counts_rhs[cls])
                                        rhs_cell.border = thin_border
                                        rhs_cell.alignment = Alignment(horizontal='center')
                                        rhs_cell.font = Font(name='Aptos Narrow', size=10)

                                    
                                    total_cell = worksheet.cell(row=row_number, column=4+len(vehicle_classes)+1, value=total_rhs)
                                    total_cell.border = thin_border
                                    total_cell.alignment = Alignment(horizontal='center')
                                    total_cell.font = Font(name='Aptos Narrow', size=10)

                                    start_row += 1 
                                    row_number += 1  

                                    
                                grand_total = sum(total_rhs_list)
                                grand_total_cell = worksheet.cell(row=row_number, column=4+len(vehicle_classes)+1, value=grand_total)
                                grand_total_cell.border = thin_border
                                grand_total_cell.alignment = Alignment(horizontal='center')
                                grand_total_cell.font = Font(name='Aptos Narrow', size=10) 

                                print(f"Grand Total RHS Count: {grand_total}")  
                                
                if lid == 'line_id_3_down':
                    print("erfeg")

                    headers = ['Date', 'Start Time', 'End Time', 
                                'Total Vehicles (UP)', *[(f'{cls} ') for cls in vehicle_class_keys],
                            
                        'Total Vehicles (DOWN)', *[(f'{cls} ') for cls in vehicle_class_keys]]
                    last_column_index = len(headers)  
                    overall_totals = [0] * len(headers) 

                    for row in range(7, start_row):
                        for col in range(4, len(headers)  + 1):
                            overall_totals[col - 4] += worksheet.cell(row=row, column=col).value or 0  # Corrected index

                    for col in range(4, len(headers)  + 1):
                        total_cell = worksheet.cell(row=start_row, column=col, value=overall_totals[col - 4])  # Corrected index
                        total_cell.border = thin_border
                        total_cell.alignment = Alignment(horizontal='center') 
                        total_cell.font = Font(name='Aptos Narrow', size=10)
                
            else:
                print("in polygon",line_id)
                if line_id == 'polygon-up':
                   

                    # Select the appropriate data source
                    data_source = polygon_1_data

                    rhs_start_col = lhs_end_col + 1  
                    rhs_end_col = rhs_start_col + len(vehicle_class_keys)

                    up_cell = worksheet.cell(row=5, column=lhs_start_col)
                    up_cell.value = "Polygon-UP"
                    up_cell.font = Font(name='Aptos Narrow', size=10, bold=True, color="00CCFF")  
                    up_cell.alignment = Alignment(horizontal='center')  
                    up_cell.border = Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)
                    

                    towards_padina_cell = worksheet.cell(row=4, column=lhs_start_col)
                    towards_padina_cell.value = "UP"
                    towards_padina_cell.font = Font(name='Aptos Narrow',size=10, bold=True, color="00CCFF")  
                    towards_padina_cell.alignment = Alignment(horizontal='center')  
                    towards_padina_cell.border=Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)

                    vehicle_class_keys = list(vehicle_classes.keys())
                    headers = ['Date', 'Start Time', 'End Time', 'Total Vehicles (UP)', *[(f'{cls} ') for cls in vehicle_class_keys]]

                    for i, header in enumerate(headers, start=1):
                        cell = worksheet.cell(row=6, column=i, value=header)
                        cell.font = Font(name='Aptos Narrow', size=10, bold=True)
                        cell.alignment = Alignment(horizontal='center', wrap_text=True) 
                        cell.border = Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)

                    last_column_index = len(headers)
                    worksheet.merge_cells(start_row=3, start_column=lhs_start_col, end_row=3, end_column=last_column_index)
                    worksheet.merge_cells(start_row=4, start_column=lhs_start_col, end_row=4, end_column=last_column_index)
                    worksheet.merge_cells(start_row=5, start_column=lhs_start_col, end_row=5, end_column=last_column_index)
                    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column_index)
                    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column_index)
                    
                    start_row = 7
                    print("hourly vehicle counts:", vehicle_counts_up)

                    if str(current_date.date()) in vehicle_counts_up:
                        hours = vehicle_counts_up[str(current_date.date())]
                        print("hours:", hours)

                        grand_total_lhs = 0  

                        for hour in range(24):
                            print(f"Processing hour: {hour}")
                            total_lhs = 0
                            counts_lhs = {cls: 0 for cls in vehicle_class_keys}

                            if line_id in data_source and str(current_date.date()) in data_source[line_id]:
                                kit_counts = data_source[line_id][str(current_date.date())].get(hour, {cls: 0 for cls in vehicle_class_keys})
                                for cls in vehicle_class_keys:
                                    counts_lhs[cls] += kit_counts.get(cls, 0)
                                    total_lhs += counts_lhs[cls]

                            start_time = f"{hour:02d}:00"
                            end_time = f"{hour + 1:02d}:00" if hour < 23 else "23:59"

                            worksheet.cell(row=start_row, column=1, value=current_date.strftime('%d%b%y').lower()).border = thin_border
                            worksheet.cell(row=start_row, column=2, value=start_time).border = thin_border
                            worksheet.cell(row=start_row, column=3, value=end_time).border = thin_border

                            total_cell = worksheet.cell(row=start_row, column=4, value=total_lhs)
                            total_cell.border = thin_border  
                            total_cell.alignment = Alignment(horizontal='center', wrap_text=True)

                            for i, cls in enumerate(vehicle_class_keys):
                                cell = worksheet.cell(row=start_row, column=4 + i + 1, value=counts_lhs[cls])
                                cell.border = thin_border
                                cell.alignment = Alignment(horizontal='center', wrap_text=True)
                                cell.font = Font(name='Aptos Narrow', size=10)

                            grand_total_lhs += total_lhs
                            start_row += 1

                        grand_total_cell = worksheet.cell(row=start_row, column=4, value=grand_total_lhs)
                        grand_total_cell.alignment = Alignment(horizontal='center', wrap_text=True)
                        grand_total_cell.border = thin_border  

                        overall_totals = [0] * len(headers)
                        for row in range(7, start_row):
                            for col in range(4, len(headers) + 1):
                                overall_totals[col - 1] += worksheet.cell(row=row, column=col).value or 0

                        for col in range(4, len(headers) + 1):
                            total_cell = worksheet.cell(row=start_row, column=col, value=overall_totals[col - 1])
                            total_cell.border = thin_border
                            total_cell.alignment = Alignment(horizontal='center') 
                            total_cell.font = Font(name='Aptos Narrow', size=10)
                            
                elif line_id =="polygon-down":
                        
                   

                    # Select the appropriate data source
                        data_source = polygon_2_data

                        rhs_start_col = lhs_end_col + 1  
                        rhs_end_col = rhs_start_col + len(vehicle_class_keys)

                        up_cell = worksheet.cell(row=5, column=lhs_start_col)
                        up_cell.value = "Polygon-DOWN"
                        up_cell.font = Font(name='Aptos Narrow', size=10, bold=True, color="00CCFF")  
                        up_cell.alignment = Alignment(horizontal='center')  
                        up_cell.border = Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)


                        towards_padina_cell = worksheet.cell(row=4, column=lhs_start_col)
                        towards_padina_cell.value = "Down"
                        towards_padina_cell.font = Font(name='Aptos Narrow',size=10, bold=True, color="00CCFF")  
                        towards_padina_cell.alignment = Alignment(horizontal='center')  
                        towards_padina_cell.border=Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)


                        vehicle_class_keys = list(vehicle_classes.keys())
                        headers = ['Date', 'Start Time', 'End Time', 'Total Vehicles (DOWN)', *[(f'{cls} ') for cls in vehicle_class_keys]]

                        for i, header in enumerate(headers, start=1):
                            cell = worksheet.cell(row=6, column=i, value=header)
                            cell.font = Font(name='Aptos Narrow', size=10, bold=True)
                            cell.alignment = Alignment(horizontal='center', wrap_text=True) 
                            cell.border = Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)

                        last_column_index = len(headers)
                        worksheet.merge_cells(start_row=3, start_column=lhs_start_col, end_row=3, end_column=last_column_index)
                        worksheet.merge_cells(start_row=4, start_column=lhs_start_col, end_row=4, end_column=last_column_index)
                        worksheet.merge_cells(start_row=5, start_column=lhs_start_col, end_row=5, end_column=last_column_index)
                        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column_index)
                        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column_index)
                        
                        start_row = 7
                        print("hourly vehicle counts:", vehicle_counts_down)

                        if str(current_date.date()) in vehicle_counts_down:
                            hours = vehicle_counts_down[str(current_date.date())]
                            print("hours:", hours)

                            grand_total_lhs = 0  

                            for hour in range(24):
                                print(f"Processing hour: {hour}")
                                total_lhs = 0
                                counts_lhs = {cls: 0 for cls in vehicle_class_keys}

                                if line_id in data_source and str(current_date.date()) in data_source[line_id]:
                                    kit_counts = data_source[line_id][str(current_date.date())].get(hour, {cls: 0 for cls in vehicle_class_keys})
                                    for cls in vehicle_class_keys:
                                        counts_lhs[cls] += kit_counts.get(cls, 0)
                                        total_lhs += counts_lhs[cls]

                                start_time = f"{hour:02d}:00"
                                end_time = f"{hour + 1:02d}:00" if hour < 23 else "23:59"

                                worksheet.cell(row=start_row, column=1, value=current_date.strftime('%d%b%y').lower()).border = thin_border
                                worksheet.cell(row=start_row, column=2, value=start_time).border = thin_border
                                worksheet.cell(row=start_row, column=3, value=end_time).border = thin_border

                                total_cell = worksheet.cell(row=start_row, column=4, value=total_lhs)
                                total_cell.border = thin_border  
                                total_cell.alignment = Alignment(horizontal='center', wrap_text=True)

                                for i, cls in enumerate(vehicle_class_keys):
                                    cell = worksheet.cell(row=start_row, column=4 + i + 1, value=counts_lhs[cls])
                                    cell.border = thin_border
                                    cell.alignment = Alignment(horizontal='center', wrap_text=True)
                                    cell.font = Font(name='Aptos Narrow', size=10)

                                grand_total_lhs += total_lhs
                                start_row += 1

                            grand_total_cell = worksheet.cell(row=start_row, column=4, value=grand_total_lhs)
                            grand_total_cell.alignment = Alignment(horizontal='center', wrap_text=True)
                            grand_total_cell.border = thin_border  

                            overall_totals = [0] * len(headers)
                            for row in range(7, start_row):
                                for col in range(4, len(headers) + 1):
                                    overall_totals[col - 1] += worksheet.cell(row=row, column=col).value or 0

                            for col in range(4, len(headers) + 1):
                                total_cell = worksheet.cell(row=start_row, column=col, value=overall_totals[col - 1])
                                total_cell.border = thin_border
                                total_cell.alignment = Alignment(horizontal='center') 
                                total_cell.font = Font(name='Aptos Narrow', size=10)
                    
                elif line_id=='polygon-up,down':
                        
                        print("polygon up and down")
                        start_row = 7
                        print("hourly vehicle counts:",vehicle_counts)

                        if str(current_date.date()) in vehicle_counts:
                            print("!!!!!!!!!!!!!!!!!!!!!!")
                            hours = vehicle_counts_up[str(current_date.date())]
                            print("hours:",hours)     


                        for hour in range(24):
                            print(f"Processing hour: {hour}")
                            total_lhs = total_rhs = total_counts = 0
                            counts_lhs = {cls: 0 for cls in vehicle_class_keys}
                            counts_rhs = {cls: 0 for cls in vehicle_class_keys}
                #
                    
                        total_start_col = lhs_end_col + 1 
                        total_end_col = total_start_col + len(vehicle_class_keys) 

                        down_cell = worksheet.cell(row=5, column=lhs_start_col)
                        down_cell.value = "UP & DOWN"
                        down_cell.font = Font(name='Aptos Narrow',size=10, bold=True, color="00CCFF")  
                        down_cell.alignment = Alignment(horizontal='center')  
                        down_cell.border=Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)

                        towards_padina_cell = worksheet.cell(row=4, column=lhs_start_col)
                        towards_padina_cell.value = f"{line_id}"
                        towards_padina_cell.font = Font(name='Aptos Narrow',size=10, bold=True, color="00CCFF")  
                        towards_padina_cell.alignment = Alignment(horizontal='center')  
                        towards_padina_cell.border=Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)

                        total_heading_cell = worksheet.cell(row=3, column=lhs_start_col)

                        vehicle_class_keys = list(vehicle_classes.keys())
                        headers = ['Date', 'Start Time', 'End Time', 
                                'Total Vehicles (UP)', *[(f'{cls} ') for cls in vehicle_class_keys],
                            
                        'Total Vehicles (DOWN)', *[(f'{cls} ') for cls in vehicle_class_keys]]

                        for i, header in enumerate(headers, start=1):
                            cell = worksheet.cell(row=6, column=i, value=header)
                            cell.font = Font(name='Aptos Narrow',size=10,bold=True)
                            cell.alignment = Alignment(horizontal='center',wrap_text=True) 
                            cell.border = Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)

                        last_column_index = len(headers) 
                        worksheet.merge_cells(start_row=3, start_column=lhs_start_col, end_row=3, end_column=last_column_index)
                        worksheet.merge_cells(start_row=4, start_column=lhs_start_col, end_row=4, end_column=last_column_index)
                        worksheet.merge_cells(start_row=5, start_column=lhs_start_col, end_row=5, end_column=4+len(vehicle_classes))

                        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column_index)
                        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column_index)

                        print("current date:", current_date)
                  
                        print("Formatted current date:", str(current_date.date()))

                        print("Type of current_date.date():", type(current_date.date()))
                        print("Type of keys in polygon_3_data['up']:", type(list(polygon_3_data['up'].keys())[0]))

                        formatted_date = str(current_date.date())

                        total_rhs_list = []  
                        row_number = 7  

                        if formatted_date in polygon_3_data['up']:
                            print(f"Processing line2 data:{polygon_3_data}")

                            for hour in range(24):  
                            
                                kit2_counts = polygon_3_data['up'][formatted_date].get(hour, {cls: 0 for cls in vehicle_class_keys})
                                counts_rhs = {cls: 0 for cls in vehicle_class_keys}  

                                for cls in vehicle_class_keys:
                                    counts_rhs[cls] = kit2_counts.get(cls, 0)

                                total_rhs = sum(counts_rhs.values())  
                                total_rhs_list.append(total_rhs)  

                                print(f"Hour: {hour}, Total RHS Count: {total_rhs}")

                                start_time = f"{hour:02d}:00"
                                end_time = f"{hour + 1:02d}:00" if hour < 23 else "23:59"

                                print("start_time:", start_time)
                                print("end_time:", end_time)

                            
                                worksheet.cell(row=start_row, column=1, value=current_date.strftime('%d%b%y').lower()).border = thin_border
                                worksheet.cell(row=start_row, column=2, value=start_time).border = thin_border
                                worksheet.cell(row=start_row, column=3, value=end_time).border = thin_border

                                for i, cls in enumerate(vehicle_class_keys):
                                    rhs_cell = worksheet.cell(row=start_row, column=4 + i + 1, value=counts_rhs[cls])
                                    rhs_cell.border = thin_border
                                    rhs_cell.alignment = Alignment(horizontal='center')
                                    rhs_cell.font = Font(name='Aptos Narrow', size=10)

                            
                                total_cell = worksheet.cell(row=row_number, column=4, value=total_rhs)
                                total_cell.border = thin_border
                                total_cell.alignment = Alignment(horizontal='center')
                                total_cell.font = Font(name='Aptos Narrow', size=10)

                                start_row += 1 
                                row_number += 1  

                            grand_total = sum(total_rhs_list)
                            grand_total_cell = worksheet.cell(row=row_number, column=4, value=grand_total)
                            grand_total_cell.border = thin_border
                            grand_total_cell.alignment = Alignment(horizontal='center')
                            grand_total_cell.font = Font(name='Aptos Narrow', size=10) 

                            print(f"Grand Total RHS Count: {grand_total}")      

                  #####################################down#################################################################################################33
                            print("polygon 3")
                            start_row = 7
                            print("hourly vehicle counts:",vehicle_counts_down)

                            if str(current_date.date()) in vehicle_counts_down:
                                print("!!!!!!!!!!!!!!!!!!!!!!")
                                hours = vehicle_counts_down[str(current_date.date())]
                                print("hours:",hours)     


                                for hour in range(24):
                                    print(f"Processing hour: {hour}")
                                    total_lhs = total_rhs = total_counts = 0
                                    counts_lhs = {cls: 0 for cls in vehicle_class_keys}
                                    counts_rhs = {cls: 0 for cls in vehicle_class_keys}
                            #
                                
                                total_start_col = lhs_end_col + 1 
                                total_end_col = total_start_col + len(vehicle_class_keys) 

                            
                                down_cell = worksheet.cell(row=5, column=lhs_start_col+len(vehicle_classes)+1)
                                down_cell.value = "Down"
                                down_cell.font = Font(name='Aptos Narrow',size=10, bold=True, color="00CCFF")  
                                down_cell.alignment = Alignment(horizontal='center')  
                                down_cell.border=Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)

                                total_heading_cell = worksheet.cell(row=3, column=lhs_start_col)

                                vehicle_class_keys = list(vehicle_classes.keys())
                                headers = ['Date', 'Start Time', 'End Time', 
                                            'Total Vehicles (UP)', *[(f'{cls} ') for cls in vehicle_class_keys],
                                        
                                    'Total Vehicles (DOWN)', *[(f'{cls} ') for cls in vehicle_class_keys]]

                                for i, header in enumerate(headers, start=1):
                                    cell = worksheet.cell(row=6, column=i, value=header)
                                    cell.font = Font(name='Aptos Narrow',size=10,bold=True)
                                    cell.alignment = Alignment(horizontal='center',wrap_text=True) 
                                    cell.border = Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)

                                last_column_index = len(headers) 
                                worksheet.merge_cells(start_row=3, start_column=lhs_start_col+len(vehicle_classes)+1, end_row=3, end_column=last_column_index)


                                worksheet.merge_cells(start_row=4, start_column=lhs_start_col+len(vehicle_classes)+1, end_row=4, end_column=last_column_index)
                                worksheet.merge_cells(start_row=5, start_column=lhs_start_col+len(vehicle_classes)+1, end_row=5, end_column=last_column_index)
                                    

                                print("current date:", current_date)
                                print("Available keys in line_3_down_data:", polygon_3_data['down'].keys())
                                print("Formatted current date:", str(current_date.date()))

                                print("Type of current_date.date():", type(current_date.date()))
                                print("Type of keys in polygon_3_data['down']:", type(list(polygon_3_data['down'].keys())[0]))

                                formatted_date = str(current_date.date())

                                total_rhs_list = []  
                                row_number = 7  

                                if formatted_date in polygon_3_data['down']:
                                    print(f"Processing down data:{polygon_3_data}")

                                    for hour in range(24):  
                                        
                                        kit2_counts = polygon_3_data['down'][formatted_date].get(hour, {cls: 0 for cls in vehicle_class_keys})
                                        counts_rhs = {cls: 0 for cls in vehicle_class_keys}  

                                        for cls in vehicle_class_keys:
                                            counts_rhs[cls] = kit2_counts.get(cls, 0)

                                        total_rhs = sum(counts_rhs.values())  
                                        total_rhs_list.append(total_rhs)  

                                        print(f"Hour: {hour}, Total RHS Count: {total_rhs}")

                                        start_time = f"{hour:02d}:00"
                                        end_time = f"{hour + 1:02d}:00" if hour < 23 else "23:59"

                                        print("start_time:", start_time)
                                        print("end_time:", end_time)

                                        
                                        worksheet.cell(row=start_row, column=1, value=current_date.strftime('%d%b%y').lower()).border = thin_border
                                        worksheet.cell(row=start_row, column=2, value=start_time).border = thin_border
                                        worksheet.cell(row=start_row, column=3, value=end_time).border = thin_border

                                        for i, cls in enumerate(vehicle_class_keys):
                                            rhs_cell = worksheet.cell(row=start_row, column=4+len(vehicle_classes) + i + 2, value=counts_rhs[cls])
                                            rhs_cell.border = thin_border
                                            rhs_cell.alignment = Alignment(horizontal='center')
                                            rhs_cell.font = Font(name='Aptos Narrow', size=10)

                                        
                                        total_cell = worksheet.cell(row=row_number, column=4+len(vehicle_classes)+1, value=total_rhs)
                                        total_cell.border = thin_border
                                        total_cell.alignment = Alignment(horizontal='center')
                                        total_cell.font = Font(name='Aptos Narrow', size=10)

                                        start_row += 1 
                                        row_number += 1  

                                        
                                    grand_total = sum(total_rhs_list)
                                    grand_total_cell = worksheet.cell(row=row_number, column=4+len(vehicle_classes)+1, value=grand_total)
                                    grand_total_cell.border = thin_border
                                    grand_total_cell.alignment = Alignment(horizontal='center')
                                    grand_total_cell.font = Font(name='Aptos Narrow', size=10) 

                                    print(f"Grand Total RHS Count: {grand_total}")  
                                    
                                    last_column_index = len(headers)  
                                    overall_totals = [0] * len(headers) 

                                    for row in range(7, start_row):
                                        for col in range(4, len(headers)  + 1):
                                            overall_totals[col - 4] += worksheet.cell(row=row, column=col).value or 0  # Corrected index

                                    for col in range(4, len(headers)  + 1):
                                        total_cell = worksheet.cell(row=start_row, column=col, value=overall_totals[col - 4])  # Corrected index
                                        total_cell.border = thin_border
                                        total_cell.alignment = Alignment(horizontal='center') 
                                        total_cell.font = Font(name='Aptos Narrow', size=10)
                          
                    
           
            worksheet['A1'].font = Font(name='Aptos Narrow', color=colors['title'], bold=True)
            worksheet['A2'].font = Font(name='Aptos Narrow', color=colors['header'], bold=True)
            worksheet['A3'].font = Font(name='Aptos Narrow', color=colors['header'], bold=True)
            worksheet['D3'].font = Font(name='Aptos Narrow', color=colors['header'], bold=True)
            worksheet['I3'].font = Font(name='Aptos Narrow', color=colors['header'], bold=True)
            worksheet['N3'].font = Font(name='Aptos Narrow', color=colors['header'], bold=True)
            worksheet['D4'].font = Font(name='Aptos Narrow', color=colors['sub_header'], bold=True,size=10)
            worksheet['I4'].font = Font(name='Aptos Narrow', color=colors['sub_header'], bold=True)
            worksheet['N4'].font = Font(name='Aptos Narrow', color=colors['sub_header'], bold=True)
            worksheet['C3'].font = Font(name='Aptos Narrow', color=colors['sub_header'], bold=True)
            worksheet['A31'].font = Font(name='Aptos Narrow', bold=True)

            worksheet.row_dimensions[1].height = 30  
            worksheet.row_dimensions[2].height = 30  
            worksheet.row_dimensions[3].height = 15
            worksheet.row_dimensions[6].height = 50

            

           
            
            for col in range(1, last_column_index + 5):  
                col_letter = get_column_letter(col)  
                print("col_letter:",col_letter)
                worksheet.column_dimensions[col_letter].width = 16

            print("@@@@#$@#%")
            

            current_date += timedelta(days=1)
            print("current :",current_date)
            