
from fastapi import FastAPI,APIRouter, HTTPException,Request
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse,StreamingResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment
from apps.main.utils.local_vendor import *
from apps.main.database.db import get_db
import zipfile
from io import BytesIO
import io,os,logging
import requests
from apps.main.config import *
from openpyxl.styles import Font, Alignment, Border, Side
from sqlalchemy import text
from apps.main.routers.super_admin.super_admin import has_permission_bool
from apps.main.utils.session import handle_session
app = FastAPI()
router=APIRouter()
logger = logging.getLogger(__name__)
templates = Jinja2Templates(directory="apps/main/front_end/templates")



#--------------------TO FETCH AVAILABLE LPU--------------------------------------------------------------
 
@router.get('/available_projects', response_class=JSONResponse)
async def get_available_project():
    try:
        db_session = next(get_db()) 
        projects = db_session.execute(text('SELECT project_id, project_name FROM project')).fetchall()
        print("NNNNNN",projects)
        return [{"project_id":project[0],"project_name": project[1]} for project in projects]
    except Exception as e:
        print(f"Error fetching kits: {e}")
        raise HTTPException(status_code=500, detail="Error fetching projects") from e
    finally:
        db_session.close()


@router.get('/drawline/available_projects/{camera_ip}', response_class=JSONResponse)
async def get_project(camera_ip: str):
    try:
        db_session = next(get_db())
       
        projects = db_session.execute(
            text('SELECT project_id, project_name FROM project WHERE camera_ip = :camera_ip'),
            {"camera_ip": camera_ip}
        ).fetchall()
        

     
        if not projects:
            return []
 
        return [{"project_id": project[0], "project_name": project[1]} for project in projects]
    except Exception as e:
        print(f"Error fetching projects for camera IP {camera_ip}: {e}")
        raise HTTPException(status_code=500, detail="Error fetching projects") from e
    finally:
        db_session.close()

@router.get('/backup/available_kits', response_class=JSONResponse)
async def get_available_kits():
    try:
        db_session = next(get_db()) 
        kits = db_session.execute(text('SELECT lpu_id, lpu_name FROM lpu_group')).fetchall()
        print("NNNNNN",kits)
        return [{"lpu_id": kit[0], "lpu_name": kit[1]} for kit in kits]
    except Exception as e:
        print(f"Error fetching kits: {e}")
        raise HTTPException(status_code=500, detail="Error fetching kits") from e
    finally:
        db_session.close()

#---------------------------REALTIME VIEW REPORT-----------------------------------------------------------

@router.get('/image/{image_name}', response_class=FileResponse)
async def get_image(image_name: str):
    file_path = image_name
    print(f"Checking for image at:{image_name} ,{file_path}")
    
    if not os.path.isfile(file_path):
        raise HTTPException(status_code=404, detail="Image not found")
    
    return FileResponse(file_path)



@router.get('/get_cameraip/{project_id}', response_class=JSONResponse)
async def get_available_cameras(project_id: int):
    try:
        db_session = next(get_db())
        
       
        cameras = db_session.execute(
            text('SELECT camera_ip FROM project WHERE project_id = :project_id'),
            {'project_id': project_id}
        ).fetchall()
        print("cameras:",cameras)
        
        return [{"camera_ip": camera[0]} for camera in cameras]
    except Exception as e:
        print(f"Error fetching cameras: {e}")
        raise HTTPException(status_code=500, detail="Error fetching cameras") from e
    finally:
        db_session.close()


import tempfile

@router.get("/main/project", response_class=HTMLResponse, name="main.project")
async def transaction_detail(request: Request):

    session_data, error_response = handle_session(request)
    if error_response:
        return RedirectResponse(url="/")
    role_id = session_data.get('role_id')

    db = next(get_db()) 

    modulee_id = 5
    show_required_permission = "show" 

    show_required_permission = await has_permission_bool(db, role_id, modulee_id, show_required_permission)
    print("show_required_permission",show_required_permission)
    if role_id == 0:
        required_permission_s = True
        edit_required_permission_s = True
        delete_required_permission_s = True
        return templates.TemplateResponse("super_admin/project/project.html", {
                                        "request": request,
                                        'page_permission':required_permission_s,
                                        'edit_permission' :edit_required_permission_s,
                                        'delete_permission' :delete_required_permission_s,                                        
                                        "session": session_data
                                        } )
    else:
        if show_required_permission:
            required_permission= "read"
            required_permission = await has_permission_bool(db, role_id, modulee_id, required_permission)
            print("required_permission",required_permission)


            edit_required_permission= "update"
            edit_required_permission = await has_permission_bool(db, role_id, modulee_id, edit_required_permission)
            print("edit_required_permission",edit_required_permission)



            delete_required_permission= "delete"
            delete_required_permission = await has_permission_bool(db, role_id, modulee_id, delete_required_permission)
            print("delete_required_permission",delete_required_permission)


            return templates.TemplateResponse("super_admin/project/project.html", {
                                            'request': request,
                                            'page_permission':required_permission,
                                            'edit_permission' :edit_required_permission,
                                            'delete_permission' :delete_required_permission,
                                            "session": session_data
                                            })
        else:
            print("Unautorize user")
            error_page = templates.get_template("error_page.html")
            content = error_page.render({"request": request})
            
            return HTMLResponse(content=content, status_code=403)




@router.get("/main/realtime/hourlyreport/", response_class=HTMLResponse, name="main.hourlyreport")
async def transaction_detail(request: Request):

    session_data, error_response = handle_session(request)
    if error_response:
        return RedirectResponse(url="/")
    role_id = session_data.get('role_id')

    db = next(get_db()) 

    modulee_id = 8
    show_required_permission = "show" 

    show_required_permission = await has_permission_bool(db, role_id, modulee_id, show_required_permission)
    print("show_required_permission",show_required_permission)
    if role_id == 0:
        required_permission_s = True
        edit_required_permission_s = True
        delete_required_permission_s = True
        return templates.TemplateResponse("report_template/realtimereport.html", {
                                        "request": request,
                                        'page_permission':required_permission_s,
                                        'edit_permission' :edit_required_permission_s,
                                        'delete_permission' :delete_required_permission_s,                                        
                                        "session": session_data
                                        } )
    else:
        if show_required_permission:
            required_permission= "read"
            required_permission = await has_permission_bool(db, role_id, modulee_id, required_permission)
            print("required_permission",required_permission)


            edit_required_permission= "update"
            edit_required_permission = await has_permission_bool(db, role_id, modulee_id, edit_required_permission)
            print("edit_required_permission",edit_required_permission)



            delete_required_permission= "delete"
            delete_required_permission = await has_permission_bool(db, role_id, modulee_id, delete_required_permission)
            print("delete_required_permission",delete_required_permission)


            return templates.TemplateResponse("report_template/realtimereport.html", {
                                            'request': request,
                                            'page_permission':required_permission,
                                            'edit_permission' :edit_required_permission,
                                            'delete_permission' :delete_required_permission,
                                            "session": session_data
                                            })
        else:
            print("Unautorize user")
            error_page = templates.get_template("error_page.html")
            content = error_page.render({"request": request})
            
            return HTMLResponse(content=content, status_code=403)






#------------------------------------------REALTIME DOWNLOAD RECORDS----------------------------------------------------------------------------------

# async def fetchdb_records_within_time_range(start_datetime: datetime, end_datetime: datetime, project_id: int, camera_ip: str, detection_id: int,line_id:str,selected_custom_class:str):

#     db_session = next(get_db())
#     all_records = []

#     print("selected_custom_class",selected_custom_class,"*********^%$^$%^")
    
#     if selected_custom_class == "1":
#         # NETC Classes → ai_class directly
#         query = """
#             SELECT id, date_time, direction, vehicle_id,
#                    vehicle_class_id, ai_class AS vehicle_name,
#                    image_path, line_id, detection_id, mapped_line
#             FROM kit1_objectdetection
#             WHERE project_id = :project_id
#               AND camera_ip = :camera_ip
#               AND detection_id = :detection_id
#               AND date_time BETWEEN :start_date AND :end_date
#             ORDER BY date_time ASC, id ASC
#         """
#         all_records = db_session.execute(
#             text(query),
#             {
#                 "project_id": project_id,
#                 "camera_ip": camera_ip,
#                 "detection_id": detection_id,
#                 "start_date": start_datetime,
#                 "end_date": end_datetime
#             }
#         ).fetchall()

#     elif selected_custom_class == "2":
#         # Grouped Classes → mapped model_class_id but use r.ai_class as vehicle_name
#         grouped_custom_classes = db_session.execute(
#             text("SELECT id FROM custom_classes WHERE id != 1")
#         ).fetchall()

#         if grouped_custom_classes:
#             custom_class_ids = [row.id for row in grouped_custom_classes]

#             mapped_class_rows = db_session.execute(
#                 text("""
#                     SELECT custom_class_id, model_class_id
#                     FROM mapping_custom_classes
#                     WHERE custom_class_id = ANY(:custom_class_ids)
#                 """),
#                 {"custom_class_ids": custom_class_ids}
#             ).fetchall()

#             mapping_dict = {}
#             for row in mapped_class_rows:
#                 mapping_dict[row.custom_class_id] = list(row.model_class_id)

#             # Fetch records for all mapped model_class_ids
#             for cid, model_ids in mapping_dict.items():
#                 if model_ids:
#                     records = db_session.execute(
#                         text("""
#                             SELECT id, date_time, direction, vehicle_id, 
#                                    vehicle_class_id, ai_class AS vehicle_name,
#                                    image_path, line_id, detection_id, mapped_line
#                             FROM kit1_objectdetection
#                             WHERE project_id = :project_id
#                               AND camera_ip = :camera_ip
#                               AND detection_id = :detection_id
#                               AND vehicle_class_id = ANY(:model_ids)
#                               AND date_time BETWEEN :start_date AND :end_date
#                             ORDER BY date_time ASC, id ASC
#                         """),
#                         {
#                             "project_id": project_id,
#                             "camera_ip": camera_ip,
#                             "detection_id": detection_id,
#                             "model_ids": model_ids,
#                             "start_date": start_datetime,
#                             "end_date": end_datetime
#                         }
#                     ).fetchall()
#                     all_records.extend(records)

#     else:
#         # Default → ai_class directly
#         query = """
#             SELECT id, date_time, direction, vehicle_id,
#                    vehicle_class_id, ai_class AS vehicle_name,
#                    image_path, line_id, detection_id, mapped_line
#             FROM kit1_objectdetection
#             WHERE project_id = :project_id
#               AND camera_ip = :camera_ip
#               AND detection_id = :detection_id
#               AND date_time BETWEEN :start_date AND :end_date
#             ORDER BY date_time ASC, id ASC
#         """
#         all_records = db_session.execute(
#             text(query),
#             {
#                 "project_id": project_id,
#                 "camera_ip": camera_ip,
#                 "detection_id": detection_id,
#                 "start_date": start_datetime,
#                 "end_date": end_datetime
#             }
#         ).fetchall()

#     print(f"Total Records Fetched: {len(all_records)}")
#     return all_records


# async def fetchdb_records_within_time_range(start_datetime: datetime, end_datetime: datetime, project_id: int, camera_ip: str, detection_id: int,line_id:str,selected_custom_class:str):
#     db_session = next(get_db())

#     print("current hour start:",start_datetime)
#     print("current hour end:",end_datetime)
    
#     if line_id and line_id.startswith("polygon-"):
#         print("in polygon")
#         table_name = "kit1_objectdetection"
#         select_fields = '''
#                 r.id, r.date_time, r.direction, r.vehicle_id, 
#                 r.vehicle_class_id, COALESCE(cc.custom_class, r.ai_class) AS ai_class,
#                 r.image_path, r.mapped_line AS line_id, r.detection_id
#             '''
#     else:
#         print("Enterrring ........................ drawline  ... query ")
#         table_name = "kit1_objectdetection"
#         # select_fields = '''
#         #         r.id, r.date_time, r.direction, r.vehicle_id, 
#         #         r.vehicle_class_id, COALESCE(cc.custom_class, r.ai_class) AS ai_class,
#         #         r.image_path, r.line_id, r.detection_id,r.mapped_line'''
#         # select_fields = '''
#         #         r.id, r.date_time, r.direction, r.vehicle_id, 
#         #         r.vehicle_class_id, r.vehicle_name,
#         #         r.image_path, r.line_id, r.detection_id,r.mapped_line'''  %%old
#         select_fields = '''
#                 r.id, r.date_time, r.direction, r.vehicle_id, 
#                 r.vehicle_class_id, r.ai_class,
#                 r.image_path, r.line_id, r.detection_id,r.mapped_line'''
  
#     # records = db_session.execute(text(f'''
#     #     SELECT {select_fields}
#     #     FROM {table_name} r
#     #     LEFT JOIN module_class_group mcg ON r.vehicle_class_id = mcg.m_classes_id
#     #     LEFT JOIN mapping_custom_classes mcc ON mcg.m_classes_id = ANY(mcc.model_class_id)
#     #     LEFT JOIN custom_classes cc ON mcc.custom_class_id = cc.id
#     #     WHERE r.date_time BETWEEN :start_date AND :end_date AND r.project_id =:project_id AND camera_ip =:camera_ip AND detection_id =:detection_id AND cc.id IS NOT NULL
#     #     ORDER BY r.id ASC
#     # '''),
#     #     {'start_date':start_datetime,'project_id':project_id,'camera_ip':camera_ip,'detection_id':detection_id,'end_date':end_datetime}
#     # ).fetchall()
#     records = db_session.execute(text(f'''
#         SELECT {select_fields}
#         FROM {table_name} r       
#         WHERE   r.date_time BETWEEN :start_date AND :end_date AND r.project_id =:project_id AND 
#                 camera_ip =:camera_ip AND detection_id =:detection_id AND r.id IS NOT NULL
        
#         ORDER BY r.date_time ASC, r.id ASC
#     '''),
#         {'start_date':start_datetime,'project_id':project_id,'camera_ip':camera_ip,'detection_id':detection_id,'end_date':end_datetime}
#     ).fetchall()

#     print("records:#$@%%#:",records)
   

#     return records     


async def fetchdb_records_within_time_range(
    start_datetime: datetime,
    end_datetime: datetime,
    project_id: int,
    camera_ip: str,
    detection_id: int,
    line_id: str,
    selected_custom_class: str
):
    db_session = next(get_db())
    all_records = []

    print("Selected Custom Classes:", selected_custom_class)

    # Convert to list of integers
    if selected_custom_class:
        selected_class_ids = [int(x) for x in selected_custom_class.split(",") if x.strip().isdigit()]
    else:
        selected_class_ids = []

    # If no selection, return empty
    if not selected_class_ids:
        return []

    # Special handling for NETC Classes (id=1)
    if 1 in selected_class_ids:
        query = """
            SELECT id, date_time, direction, vehicle_id,
                   vehicle_class_id, ai_class AS vehicle_name,
                   image_path, line_id, detection_id, mapped_line
            FROM kit1_objectdetection
            WHERE project_id = :project_id
              AND camera_ip = :camera_ip
              AND detection_id = :detection_id
              AND date_time BETWEEN :start_date AND :end_date
            ORDER BY date_time ASC, id ASC
        """
        records = db_session.execute(
            text(query),
            {
                "project_id": project_id,
                "camera_ip": camera_ip,
                "detection_id": detection_id,
                "start_date": start_datetime,
                "end_date": end_datetime
            }
        ).fetchall()
        all_records.extend(records)
        # Remove 1 from selected_class_ids so it’s not processed again
        selected_class_ids = [cid for cid in selected_class_ids if cid != 1]

    # Fetch model_class_id lists for remaining selected custom classes
    if selected_class_ids:
        mapping_rows = db_session.execute(
            text("""
                SELECT custom_class_id, model_class_id
                FROM mapping_custom_classes
                WHERE custom_class_id = ANY(:selected_ids)
            """),
            {"selected_ids": selected_class_ids}
        ).fetchall()

        # Flatten all model_class_ids
        model_ids = []
        for row in mapping_rows:
            if row.model_class_id:
                model_ids.extend(list(row.model_class_id))

        if model_ids:
            # Fetch records matching any model_class_id
            query = """
                SELECT id, date_time, direction, vehicle_id,
                       vehicle_class_id, ai_class AS vehicle_name,
                       image_path, line_id, detection_id, mapped_line
                FROM kit1_objectdetection
                WHERE project_id = :project_id
                  AND camera_ip = :camera_ip
                  AND detection_id = :detection_id
                  AND vehicle_class_id = ANY(:model_ids)
                  AND date_time BETWEEN :start_date AND :end_date
                ORDER BY date_time ASC, id ASC
            """
            records = db_session.execute(
                text(query),
                {
                    "project_id": project_id,
                    "camera_ip": camera_ip,
                    "detection_id": detection_id,
                    "model_ids": model_ids,
                    "start_date": start_datetime,
                    "end_date": end_datetime
                }
            ).fetchall()
            all_records.extend(records)

    print(f"Total Records Fetched: {len(all_records)}")
    return all_records


def create_zip_from_excel_files(temp_dir: str) -> io.BytesIO:
    """Create a ZIP file from all the Excel files in the temp directory and return it as a BytesIO object."""
    zip_buffer = io.BytesIO()
    print("&**&")
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for filename in os.listdir(temp_dir):
            file_path = os.path.join(temp_dir, filename)
            zipf.write(file_path, os.path.basename(file_path))

    zip_buffer.seek(0)
    return zip_buffer


from datetime import datetime, timedelta

@router.get("/realtime/download_records", response_class=StreamingResponse)
async def download_records(request: Request):
    camera_ip = request.query_params.get('camera_ip')
    project_id = request.query_params.get('project_id')
    detection_id = request.query_params.get('detection_id')
    start_time_str = request.query_params.get('start_time')
    end_time_str = request.query_params.get('end_time')
    server_ip = request.query_params.get('ip')
    line_id = request.query_params.get('line_id')
    selected_custom_class = request.query_params.get('selected_custom_class')    

    if line_id:
        line_id = line_id.replace("{", "").replace("}", "")

    if selected_custom_class:
        selected_class_ids = [int(x) for x in selected_custom_class.split(",") if x.strip().isdigit()]
    else:
        selected_class_ids = []

    print("Selected custom class IDs:", selected_class_ids)
    print("#####line:", line_id)
    print("project_id:", project_id, camera_ip, detection_id, start_time_str, end_time_str, server_ip, line_id,selected_custom_class)
    
    db = next(get_db())
    projects_row = db.query(Project).filter(Project.project_id == project_id).first()
    project_name = projects_row.project_name if projects_row else None


    bold_side = Side(style='thin')
    thin_border = Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)

    first_entry = (
        db.query(Detection_log)
        .filter(Detection_log.project_id == project_id,
                Detection_log.camera_ip == camera_ip,
                Detection_log.detection_id == detection_id)
        .order_by(Detection_log.start_time.asc())
        .first()
    )

    last_entry = (
        db.query(Detection_log)
        .filter(Detection_log.project_id == project_id,
                Detection_log.camera_ip == camera_ip,
                Detection_log.detection_id == detection_id)
        .order_by(Detection_log.end_time.asc())
        .first()
    )

    start_date = first_entry.start_time if first_entry else None
    end_date = last_entry.end_time if last_entry else datetime.now()
    if not end_date:
        end_date = datetime.now()


    start_time_str = datetime.strptime(start_time_str, "%Y-%m-%d %I:%M %p")
    end_time_str = datetime.strptime(end_time_str, "%Y-%m-%d %I:%M %p")

    print("Start Date:", start_date,start_time_str)
    # end_time_str = end_time_str + timedelta(seconds=59)
    end_time_str = start_time_str.replace(minute=59, second=59, microsecond=0)
    print("2222End Date:", end_date,end_time_str)


    temp_dir = "temp_reports"
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)

    try:
        records = await fetchdb_records_within_time_range(
            start_time_str, end_time_str, project_id, camera_ip, detection_id, line_id,selected_custom_class
        )

        if not records:
            print("No records found")
            return StreamingResponse(iter([]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        print("****************************************1",records)
        wb = Workbook()
        default_sheet = wb.active
        if default_sheet.title == "Sheet" and len(wb.sheetnames) == 1:
            wb.remove(default_sheet)

        sheet_data = {}
        print("kisssssssssssssssssssssss-----",line_id)
        if line_id and line_id.startswith("line_"):
            for record in records:
                line_id = record.line_id
                mapped_line = record.mapped_line or ""

                if line_id in {"line_id_1", "line_id_2"}:
                    key = line_id
                    # print("key:#", key)
                elif "line_id_3" in mapped_line:
                    key = "line_id_3"
                    # print("key:##$$%$#%$%", key)
                else:
                    continue 
                sheet_data.setdefault(key, []).append(record)
        else:
            for record in records:
                direction = record.direction
                key = 'up' if direction == 'North' else 'down'
                sheet_data.setdefault(key, []).append(record)
        

        for line_id, data in sheet_data.items():
        
            # ws = wb[line_id] if line_id in wb.sheetnames else wb.create_sheet(title=line_id)

             # Convert line_id into readable sheet name
            sheet_name_tab = str(line_id)  # default fallback
            if str(line_id) == "1" or str(line_id) == "line_id_1":
                sheet_name_tab = "UP"
            elif str(line_id) == "2" or str(line_id) == "line_id_2":
                sheet_name_tab = "DOWN"
            else:
                sheet_name_tab = "BOTH"  # keep original if not 1/2

            # Create or get sheet
            ws = wb[sheet_name_tab] if sheet_name_tab in wb.sheetnames else wb.create_sheet(title=sheet_name_tab)


            

            heading = f"ATCC REPORT - {line_id}"
            ws.merge_cells('A1:G1')
            heading_cell = ws['A1']
            heading_cell.value = heading
            heading_cell.font = Font(size=12, name='Aptos Narrow', bold=True, color="660066")
            heading_cell.alignment = Alignment(horizontal='center')
            heading_cell.border = thin_border

            start_time_cell = ws['A2']
            start_time_cell.value = f"Date: {start_time_str}-{end_time_str}"
            start_time_cell.font = Font(size=12, name='Aptos Narrow', bold=True, color="FF0000")
            start_time_cell.alignment = Alignment(horizontal='center')
            start_time_cell.border = thin_border

            ws.merge_cells('A2:G2')

            headers = ["Project Name", "Line ID", "Vehicle ID", "Date Time", "Direction", "Vehicle Name", "Image"]
            ws.append(headers)

            column_widths = {"A": 25, "B": 15, "C": 15, "D": 30, "E": 20, "F": 20, "G": 15}
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            for col_num, headers in enumerate(headers, 1):
                cell = ws.cell(row=ws.max_row, column=col_num)
                cell.border = thin_border
                cell.font =Font(size=10, name='Aptos Narrow',bold=True)

            for row in range(1, ws.max_row + 1):
                ws.row_dimensions[row].height = 25

            image_placeholder = f'http://{server_ip}/movable/No_Image_found.jpg'

            for record in data:
                image_path = record.image_path
                # print("image path in download:",image_path)
                detection_id_path_part = "/".join(image_path.split("/")[-3:]) if image_path else ""
                if line_id and line_id.startswith("line_"):
                    image_url = f'http://{server_ip}/movable/ATCC/{project_id}/{camera_ip}/{detection_id}/{detection_id_path_part}'
                    
                    # direction = 'DOWN' if record.direction == 0 else 'UP'
                    # print("hiiiiiiiiiiiiiiiiiiiiiiiiiiiiii",record.direction,direction,type(record.direction))

                    direction = "DOWN" if int(record.direction) == 0 else "UP"
                    # print("hiiiiiiiiiiiiiiiiiiiiiiiiiiiiii", record.direction, direction)

                    
                else:
                    # print("noooooooooooooooooooooooooooooooooooooo",record)
                    direction = os.path.basename(os.path.dirname(image_path)) 
                    # print("sddddddddddddddddd",direction)
                    filename = os.path.basename(image_path) 
                    image_url=f'http://{server_ip}/movable/ATCC/{detection_id}/roi_cropped/{direction}/{filename}'
                    # print("image path inside download:",image_url)
                formatted_datetime = record.date_time.strftime("%B %d, %Y, %I:%M:%S %p") if isinstance(record.date_time, datetime) else "Invalid Date"

                
                # temp_image_file = tempfile.NamedTemporaryFile(suffix=".jpg", delete=False)
                # try:
                #     response = requests.get(image_url, stream=True)
                #     if response.status_code == 200:
                #         temp_image_file.write(response.content)
                #         image_path = temp_image_file.name

                #     else:
                #         image_path = image_placeholder
                #         # print("image path###:",image_path)
                # except requests.RequestException:
                #     image_path = image_placeholder
                # finally:
                #     temp_image_file.close()


                temp_image_file = tempfile.NamedTemporaryFile(suffix=".jpg", delete=False)
                try:
                    response = requests.get(image_url, stream=True, timeout=5)  # 👈 timeout added
                    if response.status_code == 200:
                        temp_image_file.write(response.content)
                        image_path = temp_image_file.name
                    else:
                        image_path = image_placeholder
                except requests.RequestException as e:
                    print(f"Image fetch failed: {e}")
                    image_path = image_placeholder
                finally:
                    temp_image_file.close()


                # print("ddddddddddddyes",direction)
                row = [project_name, line_id, record[3], formatted_datetime, direction, record[5]]
                ws.append(row)

                ws.row_dimensions[ws.max_row].height = 25
                for col in range(1, len(row) + 2):
                    cell = ws.cell(row=ws.max_row, column=col)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                    cell.font = Font(size=10, name='Aptos Narrow')

                if os.path.isfile(image_path):

                    # print("$%^%&%")
                    img = ExcelImage(image_path)
                    img.width = 30
                    img.height = 30
                    ws.add_image(img, f"G{ws.max_row}")

                    # print("Image inserted into Excel:", image_path)
                else:
                    print("Image file not found:", image_path)


        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        response = StreamingResponse(excel_buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response.headers["Content-Disposition"] = f"attachment; filename=records_{start_time_str}_to_{end_time_str}.xlsx"

        return response

    except Exception as e:
        # shutil.rmtree(temp_dir)
        raise e






# @router.get("/realtime/download_records", response_class=StreamingResponse)
# async def download_records(request: Request):
#     camera_ip = request.query_params.get('camera_ip')
#     project_id = request.query_params.get('project_id')
#     detection_id = request.query_params.get('detection_id')
#     start_time_str = request.query_params.get('start_time')
#     end_time_str = request.query_params.get('end_time')
#     server_ip = request.query_params.get('ip')
#     line_id = request.query_params.get('line_id')

#     if line_id:
#         line_id = line_id.replace("{", "").replace("}", "")

#     print("#####line:", line_id)
#     print("project_id:", project_id, camera_ip, detection_id, start_time_str, end_time_str, server_ip, line_id)

#     db = next(get_db())

#     bold_side = Side(style='thin')
#     thin_border = Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)

#     first_entry = (
#         db.query(Detection_log)
#         .filter(Detection_log.project_id == project_id,
#                 Detection_log.camera_ip == camera_ip,
#                 Detection_log.detection_id == detection_id)
#         .order_by(Detection_log.start_time.asc())
#         .first()
#     )

#     last_entry = (
#         db.query(Detection_log)
#         .filter(Detection_log.project_id == project_id,
#                 Detection_log.camera_ip == camera_ip,
#                 Detection_log.detection_id == detection_id)
#         .order_by(Detection_log.end_time.asc())
#         .first()
#     )

#     start_date = first_entry.start_time if first_entry else None
#     end_date = last_entry.end_time if last_entry else datetime.now()
#     if not end_date:
#         end_date = datetime.now()


#     start_time_str = datetime.strptime(start_time_str, "%Y-%m-%d %I:%M %p")
#     end_time_str = datetime.strptime(end_time_str, "%Y-%m-%d %I:%M %p")

#     print("Start Date:", start_date,start_time_str)
#     print("End Date:", end_date,end_time_str)

#     temp_dir = "temp_reports"
#     if not os.path.exists(temp_dir):
#         os.makedirs(temp_dir)

#     try:
#         records = await fetchdb_records_within_time_range(
#             start_time_str, end_time_str, project_id, camera_ip, detection_id, line_id
#         )

#         if not records:
#             print("No records found")
#             return StreamingResponse(iter([]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


#         wb = Workbook()
#         default_sheet = wb.active
#         if default_sheet.title == "Sheet" and len(wb.sheetnames) == 1:
#             wb.remove(default_sheet)

#         sheet_data = {}
        
#         if line_id and line_id.startswith("line_"):
#             for record in records:
#                 line_id = record.line_id
#                 mapped_line = record.mapped_line or ""

#                 if line_id in {"line_id_1", "line_id_2"}:
#                     key = line_id
#                     # print("key:#", key)
#                 elif "line_id_3" in mapped_line:
#                     key = "line_id_3"
#                     # print("key:##$$%$#%$%", key)
#                 else:
#                     continue 
#                 sheet_data.setdefault(key, []).append(record)
#         else:
#             for record in records:
#                 direction = record.direction
#                 key = 'up' if direction == 'North' else 'down'
#                 sheet_data.setdefault(key, []).append(record)
        

#         for line_id, data in sheet_data.items():
#             ws = wb[line_id] if line_id in wb.sheetnames else wb.create_sheet(title=line_id)

#             heading = f"ATCC REPORT - {line_id}"
#             ws.merge_cells('A1:H1')
#             heading_cell = ws['A1']
#             heading_cell.value = heading
#             heading_cell.font = Font(size=12, name='Aptos Narrow', bold=True, color="660066")
#             heading_cell.alignment = Alignment(horizontal='center')
#             heading_cell.border = thin_border

#             start_time_cell = ws['A2']
#             start_time_cell.value = f"Date: {start_time_str}-{end_time_str}"
#             start_time_cell.font = Font(size=12, name='Aptos Narrow', bold=True, color="FF0000")
#             start_time_cell.alignment = Alignment(horizontal='center')
#             start_time_cell.border = thin_border

#             ws.merge_cells('A2:H2')

#             headers = ["Project ID", "Detection ID", "Line ID", "Vehicle ID", "Date Time", "Direction", "Vehicle Name", "Image"]
#             ws.append(headers)

#             column_widths = {"A": 20, "B": 20, "C": 15, "D": 15, "E": 30, "F": 20, "G": 15, "H": 20}
#             for col, width in column_widths.items():
#                 ws.column_dimensions[col].width = width

#             for col_num, headers in enumerate(headers, 1):
#                 cell = ws.cell(row=ws.max_row, column=col_num)
#                 cell.border = thin_border
#                 cell.font =Font(size=10, name='Aptos Narrow',bold=True)

#             for row in range(1, ws.max_row + 1):
#                 ws.row_dimensions[row].height = 25

#             image_placeholder = f'http://{server_ip}/movable/No_Image_found.jpg'

#             for record in data:
#                 image_path = record.image_path
#                 # print("image path in download:",image_path)
#                 detection_id_path_part = "/".join(image_path.split("/")[-3:]) if image_path else ""
#                 if line_id and line_id.startswith("line_"):
#                     image_url = f'http://{server_ip}/movable/ATCC/{project_id}/{camera_ip}/{detection_id}/{detection_id_path_part}'

#                     direction = 'DOWN' if record.direction == 0 else 'UP'
                    
#                 else:
#                     direction = os.path.basename(os.path.dirname(image_path)) 
#                     filename = os.path.basename(image_path) 
#                     image_url=f'http://{server_ip}/movable/ATCC/{detection_id}/roi_cropped/{direction}/{filename}'
#                     # print("image path inside download:",image_url)
#                 formatted_datetime = record.date_time.strftime("%B %d, %Y, %I:%M:%S %p") if isinstance(record.date_time, datetime) else "Invalid Date"

                
#                 temp_image_file = tempfile.NamedTemporaryFile(suffix=".jpg", delete=False)
#                 try:
#                     response = requests.get(image_url, stream=True)
#                     if response.status_code == 200:
#                         temp_image_file.write(response.content)
#                         image_path = temp_image_file.name

#                     else:
#                         image_path = image_placeholder
#                         # print("image path###:",image_path)
#                 except requests.RequestException:
#                     image_path = image_placeholder
#                 finally:
#                     temp_image_file.close()

#                 row = [project_id, detection_id, line_id, record[3], formatted_datetime, direction, record[5]]
#                 ws.append(row)

#                 ws.row_dimensions[ws.max_row].height = 25
#                 for col in range(1, len(row) + 2):
#                     cell = ws.cell(row=ws.max_row, column=col)
#                     cell.alignment = Alignment(horizontal='center', vertical='center')
#                     cell.border = thin_border
#                     cell.font = Font(size=10, name='Aptos Narrow')

#                 if os.path.isfile(image_path):

#                     # print("$%^%&%")
#                     img = ExcelImage(image_path)
#                     img.width = 30
#                     img.height = 30
#                     ws.add_image(img, f"H{ws.max_row}")

#                     # print("Image inserted into Excel:", image_path)
#                 else:
#                     print("Image file not found:", image_path)


#         excel_buffer = io.BytesIO()
#         wb.save(excel_buffer)
#         excel_buffer.seek(0)

#         response = StreamingResponse(excel_buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#         response.headers["Content-Disposition"] = f"attachment; filename=records_{start_time_str}_to_{end_time_str}.xlsx"

#         return response

#     except Exception as e:
#         # shutil.rmtree(temp_dir)
#         raise e



